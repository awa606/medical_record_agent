"""Excel 模板读写工具。

代码里保留较详细的中文注释，是为了让课程项目答辩时能清楚说明：
1. 模板如何生成；
2. 字段说明和示例数据如何分离；
3. Excel 如何稳定转换为 JSON 知识库。
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.data_builder.paths import TEMPLATE_DIR, ensure_data_builder_dirs
from app.data_builder.template_data import TemplateSpec


DESCRIPTION_SHEET = "字段说明"
DATA_SHEET = "示例数据"


def _stringify_cell(value: Any) -> Any:
    """把复杂对象转换成 Excel 友好的字符串。

    Excel 单元格不适合直接保存 Python 的 list/dict。这里统一把 list 用
    竖线连接，把 dict 用 JSON 字符串保存，后续 build_kb.py 再做反向解析。
    """

    if isinstance(value, list):
        return "|".join(str(item) for item in value)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False)
    return value


def _style_sheet(ws) -> None:
    """给模板加上基础样式，让字段含义更容易阅读。"""

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    for column_cells in ws.columns:
        max_length = 8
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, min(len(str(cell.value)), 50))
        ws.column_dimensions[column_letter].width = max_length + 4


def write_template_workbook(spec: TemplateSpec, target_dir: Path = TEMPLATE_DIR) -> Path:
    """根据模板定义写出一个 Excel 文件。"""

    ensure_data_builder_dirs()
    target_path = target_dir / spec.filename

    wb = Workbook()
    desc_ws = wb.active
    desc_ws.title = DESCRIPTION_SHEET
    desc_ws.append(["字段名", "字段说明", "是否必填", "示例"])
    for item in spec.fields:
        desc_ws.append([item["字段名"], item["字段说明"], item["是否必填"], item["示例"]])
    _style_sheet(desc_ws)

    data_ws = wb.create_sheet(DATA_SHEET)
    headers = [item["字段名"] for item in spec.fields]
    data_ws.append(headers)
    for row in spec.rows:
        data_ws.append([_stringify_cell(row.get(header, "")) for header in headers])
    _style_sheet(data_ws)

    wb.properties.title = spec.title
    wb.properties.subject = "感冒中医证候-症状-问诊路径知识库模板"
    wb.properties.keywords = "模拟数据,中医证候,问诊路径,课程项目"
    wb.save(target_path)
    return target_path


def load_data_rows(workbook_path: Path) -> list[dict[str, Any]]:
    """读取 Excel 的“示例数据”工作表，返回字典列表。"""

    wb = load_workbook(workbook_path, data_only=True)
    if DATA_SHEET not in wb.sheetnames:
        raise ValueError(f"{workbook_path.name} 缺少工作表：{DATA_SHEET}")

    ws = wb[DATA_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(cell).strip() for cell in rows[0] if cell is not None]
    records: list[dict[str, Any]] = []
    for row_values in rows[1:]:
        record: dict[str, Any] = {}
        for index, header in enumerate(headers):
            value = row_values[index] if index < len(row_values) else None
            if isinstance(value, str):
                value = value.strip()
            record[header] = value

        # 整行为空时跳过，方便老师或同学在 Excel 里预留空行。
        if any(value not in (None, "") for value in record.values()):
            records.append(record)
    return records
